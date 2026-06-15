# services/excel_service.py
import os
import openpyxl
import pandas as pd
from datetime import datetime
from openpyxl.chart import PieChart, Reference

class ExcelService:
    def __init__(self, file="news_log.xlsx"):
        self.file = file
        self.titles = set()
        self._load_existing_titles()


    def _load_existing_titles(self):
        if not os.path.exists(self.file):
            return
        try:
            df = pd.read_excel(self.file, sheet_name="News Log")
            if "ニュースタイトル" in df.columns:
                self.titles = set(df["ニュースタイトル"].astype(str).str.strip())
        except Exception:
            pass


    def is_duplicate(self, title: str) -> bool:
        return title.strip() in self.titles


    def write(self, category, title, url, summary, importance):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = [now, category, title, url, summary or "要約なし", importance]

        if os.path.exists(self.file):
            wb = openpyxl.load_workbook(self.file)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "News Log"
            ws.append([
                "通知日時",
                "カテゴリ",
                "ニュースタイトル",
                "記事URL",
                "AI要約内容",
                "重要度スコア"
            ])

        ws.append(row)
        wb.save(self.file)

        self.titles.add(title.strip())
        self.update_chart()


    def update_chart(self):
        if not os.path.exists(self.file):
            return

        df = pd.read_excel(self.file)
        if df.empty or "カテゴリ" not in df.columns:
            return

        summary = df["カテゴリ"].value_counts().reset_index()
        summary.columns = ["カテゴリ", "件数"]

        wb = openpyxl.load_workbook(self.file)
        name = "集計レポート"

        if name in wb.sheetnames:
            del wb[name]

        ws = wb.create_sheet(name)
        ws.append(["カテゴリ", "件数"])

        for _, row in summary.iterrows():
            ws.append([row["カテゴリ"], row["件数"]])

        pie = PieChart()
        pie.title = "カテゴリ比率"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(summary) + 1)
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(summary) + 1)

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, "D2")

        wb.save(self.file)
