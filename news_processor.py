import os
import requests
from datetime import datetime
from bs4 import BeautifulSoup
from google import genai
import openpyxl
from openpyxl.chart import PieChart, Reference
import pandas as pd
from utils.http import get_text, get_client, HTTPError

# === 設定エリア ===
CONFIG_FILE = "config.txt"
EXCEL_FILE = "news_log.xlsx"
# ==================


def load_config():
    """config.txtから各種設定を読み込む"""
    config = {
        "webhook_url": None,
        "categories": ["domestic"],
        "keywords": [],
        "gemini_api_key": None,
        "ai_summary_enabled": True,
        "check_interval": 60,
        "semantic_interest": None,
        "semantic_threshold": 80,
    }

    if not os.path.exists(CONFIG_FILE):
        print(f"❌ エラー: 設定ファイル '{CONFIG_FILE}' が見つかりません。")
        return None

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue

                if line.startswith("DISCORD_WEBHOOK_URL="):
                    config["webhook_url"] = line.split("=", 1)[1].strip()
                elif line.startswith("CATEGORY="):
                    val = line.split("=", 1)[1].strip()
                    if val:
                        # カンマ区切りで複数カテゴリを受け取る
                        cats = [c.strip() for c in val.split(",") if c.strip()]
                        if cats:
                            config["categories"] = cats
                elif line.startswith("KEYWORDS="):
                    val = line.split("=", 1)[1].strip()
                    if val:
                        config["keywords"] = [k.strip() for k in val.split(",") if k.strip()]
                elif line.startswith("GEMINI_API_KEY="):
                    val = line.split("=", 1)[1].strip()
                    if val:
                        config["gemini_api_key"] = val
                elif line.startswith("AI_SUMMARY_ENABLED="):
                    val = line.split("=", 1)[1].strip()
                    if val:
                        config["ai_summary_enabled"] = val.lower() in {"1", "true", "yes", "on"}
                    else:
                        config["ai_summary_enabled"] = False
                elif line.startswith("SEMANTIC_INTEREST="):
                    val = line.split("=", 1)[1].strip()
                    if val:
                        config["semantic_interest"] = val
                elif line.startswith("SEMANTIC_THRESHOLD="):
                    val = line.split("=", 1)[1].strip()
                    if val and val.isdigit():
                        threshold = int(val)
                        if 0 <= threshold <= 100:
                            config["semantic_threshold"] = threshold
                        else:
                            print("⚠️ SEMANTIC_THRESHOLD は 0〜100 の範囲で指定してください。デフォルト 80 を使用します。")
                elif line.startswith("CHECK_INTERVAL="):
                    val = line.split("=", 1)[1].strip()
                    if val and val.isdigit():
                        config["check_interval"] = int(val)

        if not config["webhook_url"]:
            print(f"❌ エラー: '{CONFIG_FILE}' 内に DISCORD_WEBHOOK_URL の値が設定されていません。")
            return None

        return config
    except Exception as e:
        print(f"❌ 設定ファイルの読み込み中にエラーが発生しました: {e}")
        return None


def get_latest_news(category):
    """指定されたカテゴリのYahoo!ニュースから最新記事を取得する"""
    base_domain = "https://news.yahoo.co.jp"
    url = f"{base_domain}/categories/{category}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    try:
        response_text = get_text(url, headers=headers)
        soup = BeautifulSoup(response_text, "html.parser")
        for link in soup.find_all("a"):
            href = link.get("href", "")
            title = link.text.strip()
            if ("pickup" in href or "articles" in href) and len(title) > 10:
                if href.startswith("/"):
                    href = f"{base_domain}{href}"
                return {"title": title, "url": href}
    except HTTPError as e:
        print(f"データ取得エラー: {e}")
    except Exception as e:
        print(f"データ取得エラー: {e}")
    return None


def is_title_already_notified(title):
    """Excelのログファイルを読み込み、同じニュースタイトルが既に記録されているかチェックする"""
    if not os.path.exists(EXCEL_FILE):
        return False  # Excelファイル自体がなければ、まだ通知されていない

    try:
        # Excelファイルを読み込む
        df = pd.read_excel(EXCEL_FILE, sheet_name="News Log")

        if df.empty or "ニュースタイトル" not in df.columns:
            return False

        # タイトル列（空白や大文字小文字を整えて比較）に同じものがあるか判定
        existing_titles = df["ニュースタイトル"].astype(str).str.strip().tolist()
        if title.strip() in existing_titles:
            return True

    except Exception as e:
        print(f"⚠️ 過去ログの重複チェック中にエラーが発生しました: {e}")

    return False


def fetch_news_body(article_url):
    """ニュースのURLから本文テキストをスクレイピングして取得する"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    try:
        body_html = get_text(article_url, headers=headers)
        soup = BeautifulSoup(body_html, "html.parser")
        paragraphs = soup.find_all("p")
        body_text = "".join([p.text.strip() for p in paragraphs if len(p.text.strip()) > 10])
        return body_text[:1500]
    except HTTPError as e:
        print(f"⚠️ 本文の取得に失敗しました: {e}")
    except Exception as e:
        print(f"⚠️ 本文の取得に失敗しました: {e}")
    return ""


def summarize_with_gemini(api_key, title, body_text):
    """Gemini APIを使用して、ニュース本文を3行の箇条書きに要約する"""
    if not api_key or not body_text:
        return None

    try:
        client = genai.Client(api_key=api_key)
        prompt = f"""
        以下のニュース記事の内容を読み、重要なポイントを3行のシンプルな箇条書き（各行の先頭は「・」）で要約してください。
        余計な挨拶や前置きは一切省き、要約文だけを出力してください。

        【タイトル】: {title}
        【本文】: {body_text}
        """
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
        )
        return response.text.strip()
    except Exception as e:
        print(f"⚠️ Gemini要約エラー: {e}")
    return None


def score_semantic_match(api_key, interest_text, title, body_text):
    """Gemini を使って、ニュースが関心分野にどの程度合致するかを 0-100 の整数スコアで返す"""
    if not api_key or not interest_text or not title or not body_text:
        return None

    try:
        client = genai.Client(api_key=api_key)
        prompt = f"""
        以下のユーザーの関心テーマとニュース記事がどの程度関連しているかを、0～100のスコアで評価してください。
        100は非常に関連性が高い、0は関連性がほとんどないことを意味します。
        数字のみを返してください。

        【ユーザーの関心】{interest_text}
        【ニュースタイトル】{title}
        【ニュース本文】{body_text}
        """
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
        )
        response_text = response.text.strip()

        import re
        match = re.search(r"\b([0-9]{1,3})\b", response_text)
        if match:
            score = int(match.group(1))
            if 0 <= score <= 100:
                return score
    except Exception as e:
        print(f"⚠️ セマンティックスコアリングエラー: {e}")

    return None


def generate_category_chart():
    """Excelのログデータからカテゴリごとの件数を集計し、円グラフを自動作成・挿入する"""
    if not os.path.exists(EXCEL_FILE):
        return

    try:
        # pandasで現在のログを読み込む
        df = pd.read_excel(EXCEL_FILE)

        # データが空、または必要な列がない場合はスキップ
        if df.empty or "カテゴリ" not in df.columns:
            return

        # カテゴリごとの件数を集計する
        summary_df = df["カテゴリ"].value_counts().reset_index()
        summary_df.columns = ["カテゴリ", "件数"]

        # openpyxlで元のExcelファイルを開く
        wb = openpyxl.load_workbook(EXCEL_FILE)

        # 「集計レポート」シートを作成（すでにあれば上書き用に一度削除して再作成）
        summary_sheet_name = "集計レポート"
        if summary_sheet_name in wb.sheetnames:
            del wb[summary_sheet_name]
        ws_report = wb.create_sheet(title=summary_sheet_name)

        # 集計したデータを書き込む
        ws_report.append(["カテゴリ", "通知件数"])
        for index, row in summary_df.iterrows():
            ws_report.append([row["カテゴリ"], row["件数"]])

        # 円グラフを作成する
        pie = PieChart()
        pie.title = "通知ニュースのカテゴリ比率"

        # データの範囲（件数の列: B1〜B[データの件数+1]）
        data_ref = Reference(
            ws_report, min_col=2, min_row=1, max_row=len(summary_df) + 1
        )
        # ラベルの範囲（カテゴリ名の列: A2〜A[データの件数+1]）
        labels_ref = Reference(
            ws_report, min_col=1, min_row=2, max_row=len(summary_df) + 1
        )

        # グラフにデータとラベルを設定
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        pie.style = 10  # グラフのデザインスタイル

        # グラフをD2セル付近に挿入
        ws_report.add_chart(pie, "D2")

        # Excelを上書き保存
        wb.save(EXCEL_FILE)
        print("📊 カテゴリ別の集計円グラフを自動更新しました！")

    except Exception as e:
        print(f"⚠️ グラフ作成中にエラーが発生しました: {e}")


def write_to_excel(category, title, url, summary):
    """ニュースデータをExcelファイルに書き込む（追記保存）"""
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    clean_summary = summary if summary else "要約なし"
    row_data = [now_str, category, title, url, clean_summary]

    try:
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "News Log"
            ws.append(["通知日時", "カテゴリ", "ニュースタイトル", "記事URL", "AI要約内容"])

        ws.append(row_data)
        wb.save(EXCEL_FILE)
        print(f"🗄️ Excelデータベースに記録しました: {title}")

        # 👇 【変更点3】データ書き込み成功の直後にグラフ更新関数を呼び出す
        generate_category_chart()

    except Exception as e:
        print(f"⚠️ Excelへの書き込みに失敗しました: {e}")


def send_to_discord(webhook_url, news, is_test=False, hit_word=None, summary=None, semantic_score=None):
    """Discordにニュースを通知する"""
    if news is None:
        content = "📢 【システム通知】ニュースデータの読み込みに失敗しました。"
    else:
        if is_test:
            prefix = "📢 【起動テスト成功】\n"
        elif hit_word:
            prefix = f"🎯 【キーワード的中: {hit_word}】\n"
        else:
            prefix = "📰 【新着ニュース】\n"

        content = f"{prefix}**{news['title']}**\n{news['url']}"
        if summary:
            content += f"\n\n> 🤖 **AIによる3行要約:**\n> {summary.replace('\n', '\n> ')}"
        if semantic_score is not None:
            content += f"\n\n> 🔎 **関連度スコア:** {semantic_score}/100"

    payload = {"content": content}

    try:
        client = get_client()
        client.post_json(webhook_url, payload)
        if news:
            print(f"✅ Discordへ通知しました: {news['title']}")
        else:
            print("✅ Discordへテスト通知を送信しました。")
    except HTTPError as e:
        print(f"❌ 通信エラーが発生しました: {e}")
    except Exception as e:
        print(f"❌ 通信エラーが発生しました: {e}")


def speak_text(text):
    """テキストを音声で再生する"""
    try:
        import pyttsx3
        engine = pyttsx3.init()
        engine.say(text)
        engine.runAndWait()
    except Exception as e:
        print(f"⚠️ 音声読み上げに失敗しました: {e}")
